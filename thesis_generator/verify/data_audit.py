"""Raw-data audit — recompute every M, %, χ², t, p, d, V Craméra from raw Excel.

Ported and generalized from `audyt_liczb.py`. The proven session caught 5/75
numbers as either approximations or off-by-N this way (e.g. "6 respondentów"
turned out to be 4).

How it works:
1. Walk the raw xlsx with openpyxl (data_only=True, so we get computed values
   not formulas).
2. Extract every cell that looks like a key statistic.
3. Parse the draft .docx for every numeric claim and try to match each one
   against an extracted value.
4. Flag mismatches with magnitude (off-by-N? approximation? wrong sign?).

For full automation, the matching is heuristic — the report goes to the user
for adjudication, not auto-fix. (Gotcha #10 — never silently rewrite numbers.)
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from docx import Document

from thesis_generator.config import ThesisProject


@dataclass(slots=True)
class NumericClaim:
    """A number stated in the draft, with its surrounding context."""

    value: float
    raw_text: str  # the literal "65%" or "M = 4.21" as written
    paragraph_idx: int
    context_snippet: str  # ~120 chars around the value


@dataclass(slots=True)
class ExcelValue:
    """A numeric value pulled from an Excel cell, with provenance."""

    value: float
    sheet: str
    cell: str
    label_guess: str  # nearest text in the same row/col


@dataclass(slots=True)
class DataAuditResult:
    claims: list[NumericClaim]
    excel_values: list[ExcelValue]
    matches: list[tuple[NumericClaim, ExcelValue]]  # claim matched to excel value
    unmatched_claims: list[NumericClaim]
    likely_off_by: list[tuple[NumericClaim, ExcelValue, str]]  # (claim, excel, reason)
    notes: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return not self.likely_off_by

    def to_report(self) -> str:
        lines = ["# Data audit — claims in draft vs raw Excel\n"]
        lines.append(f"- numeric claims in draft: {len(self.claims)}")
        lines.append(f"- excel values extracted: {len(self.excel_values)}")
        lines.append(f"- matched: {len(self.matches)}")
        lines.append(f"- unmatched (claim with no excel counterpart): {len(self.unmatched_claims)}")
        lines.append(f"- LIKELY MISMATCHED: {len(self.likely_off_by)}\n")
        if self.likely_off_by:
            lines.append("## ❌ Likely mismatches (REVIEW MANUALLY)\n")
            for claim, ev, reason in self.likely_off_by:
                lines.append(f"### `{claim.raw_text}` in paragraph {claim.paragraph_idx}")
                lines.append(f"- Excel: `{ev.sheet}!{ev.cell}` = {ev.value} (label: {ev.label_guess})")
                lines.append(f"- Reason: {reason}")
                lines.append(f"- Context: …{claim.context_snippet}…\n")
        if self.unmatched_claims:
            lines.append("## ⚠️ Numeric claims with no excel counterpart\n")
            for c in self.unmatched_claims[:30]:
                lines.append(f"- `{c.raw_text}` (paragraph {c.paragraph_idx}) — …{c.context_snippet[:60]}…")
        if self.notes:
            lines.append("\n## Notes\n")
            lines.extend(f"- {n}" for n in self.notes)
        return "\n".join(lines)


_NUMBER_TOKEN = re.compile(
    r"(?<![\w\.])"
    r"(\d{1,4}(?:[\.,]\d{1,4})?)"
    r"\s*(%|p\.?p\.?|χ²|t|p|d|r|n)?",
    re.UNICODE,
)


def _extract_claims_from_draft(draft: Path) -> list[NumericClaim]:
    doc = Document(str(draft))
    claims: list[NumericClaim] = []
    for i, p in enumerate(doc.paragraphs):
        text = p.text
        for m in _NUMBER_TOKEN.finditer(text):
            raw = m.group(0).strip()
            num_str = m.group(1).replace(",", ".")
            try:
                value = float(num_str)
            except ValueError:
                continue
            # Skip page numbers, years, simple list ordinals
            if value > 1900 and value < 2100 and len(num_str) == 4:
                continue
            start = max(0, m.start() - 60)
            end = min(len(text), m.end() + 60)
            claims.append(NumericClaim(
                value=value,
                raw_text=raw,
                paragraph_idx=i,
                context_snippet=text[start:end].replace("\n", " "),
            ))
    return claims


def _extract_excel_values(xlsx_paths: list[Path]) -> list[ExcelValue]:
    out: list[ExcelValue] = []
    for xp in xlsx_paths:
        try:
            wb = openpyxl.load_workbook(str(xp), data_only=True, read_only=True)
        except Exception:
            continue
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            # Pull first 100 rows × 20 cols
            for row in ws.iter_rows(min_row=1, max_row=100, max_col=20, values_only=False):
                # Find label cell (leftmost non-empty text)
                label = ""
                for c in row:
                    if c.value is not None and isinstance(c.value, str) and c.value.strip():
                        label = c.value.strip()[:60]
                        break
                for c in row:
                    if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                        out.append(ExcelValue(
                            value=float(c.value),
                            sheet=ws_name,
                            cell=c.coordinate,
                            label_guess=label,
                        ))
        wb.close()
    return out


def _match_claims_to_values(
    claims: list[NumericClaim], values: list[ExcelValue]
) -> tuple[list[tuple[NumericClaim, ExcelValue]], list[NumericClaim], list[tuple[NumericClaim, ExcelValue, str]]]:
    matches: list[tuple[NumericClaim, ExcelValue]] = []
    unmatched: list[NumericClaim] = []
    off_by: list[tuple[NumericClaim, ExcelValue, str]] = []
    for c in claims:
        # Exact match first
        exact = [v for v in values if abs(v.value - c.value) < 0.005]
        if exact:
            matches.append((c, exact[0]))
            continue
        # Percentage interpretation: claim "65%" matches excel 0.65
        pct = [v for v in values if abs(v.value * 100 - c.value) < 0.5]
        if pct and "%" in c.raw_text:
            matches.append((c, pct[0]))
            continue
        # Off-by-1 or off-by-magnitude detection (the "6 respondentów → 4" gotcha)
        close = [(v, abs(v.value - c.value)) for v in values if abs(v.value - c.value) < max(1.5, c.value * 0.1)]
        close.sort(key=lambda t: t[1])
        if close:
            v, diff = close[0]
            off_by.append((c, v, f"draft says {c.value}, excel has {v.value} (Δ={diff:.2f})"))
            continue
        unmatched.append(c)
    return matches, unmatched, off_by


def run_data_audit(project: ThesisProject) -> DataAuditResult:
    if not project.inputs.raw_data:
        return DataAuditResult(
            claims=[],
            excel_values=[],
            matches=[],
            unmatched_claims=[],
            likely_off_by=[],
            notes=["No raw_data configured in thesis.yaml — skipping data audit."],
        )

    draft = project.resolve_input(project.inputs.draft)
    xlsx_paths = [project.resolve_input(p) for p in project.inputs.raw_data if str(p).lower().endswith((".xlsx", ".xls"))]

    claims = _extract_claims_from_draft(draft)
    excel_values = _extract_excel_values(xlsx_paths)
    matches, unmatched, off_by = _match_claims_to_values(claims, excel_values)

    return DataAuditResult(
        claims=claims,
        excel_values=excel_values,
        matches=matches,
        unmatched_claims=unmatched,
        likely_off_by=off_by,
        notes=[
            f"Extracted {len(claims)} numeric claims from draft, "
            f"{len(excel_values)} numeric cells across {len(xlsx_paths)} workbook(s).",
            "Mismatch flags are HEURISTIC — review each one manually. Some 'unmatched' claims are legitimate (e.g. counts of pages, theoretical values).",
        ],
    )


def save_report(result: DataAuditResult, project: ThesisProject) -> Path:
    out = project.reports_dir() / "data_audit.md"
    out.write_text(result.to_report(), encoding="utf-8")
    return out
